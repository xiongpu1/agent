import argparse
import difflib
import os
import re
import sys
import zipfile
from posixpath import normpath
from xml.etree import ElementTree as ET
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook

# Ensure backend root (the parent of 'src') is on sys.path so this script can be
# executed from either repo root or backend/.
BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from src.dataclass import Neo4jConfig
from src.neo4j_file_add_neo4j import get_neo4j_driver


DEFAULT_XLSX_PATH = str(BACKEND_ROOT / "data_test" / "Bellagio Product Catalog V2.xlsx")
DEFAULT_SHEETS = [
    "Hot Tub大缸",
    "Swim Spa泳池",
    "Vinterkold冰水缸",
]


def _norm_match_key(s: str) -> str:
    s = (s or "").strip().lower()
    if not s:
        return ""
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"[\-_]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _similarity(a: str, b: str) -> float:
    aa = _norm_match_key(a)
    bb = _norm_match_key(b)
    if not aa or not bb:
        return 0.0
    return difflib.SequenceMatcher(None, aa, bb).ratio()


def load_all_material_codes() -> List[Tuple[str, str]]:
    neo4j_config = get_neo4j_config()
    driver = get_neo4j_driver(neo4j_config)
    out: List[Tuple[str, str]] = []
    try:
        with driver.session() as session:
            res = session.run(
                """
                MATCH (m:Material)
                WHERE m.material_code IS NOT NULL
                RETURN elementId(m) AS eid, m.material_code AS code
                """
            )
            for r in res:
                eid = r.get("eid")
                code = r.get("code")
                if eid is None or code is None:
                    continue
                out.append((str(eid), str(code)))
    finally:
        try:
            driver.close()
        except Exception:
            pass
    return out


def update_material_images_as_nodes_by_eid(
    *,
    rows: List[Dict[str, Any]],
    batch_size: int,
) -> Dict[str, int]:
    neo4j_config = get_neo4j_config()
    driver = get_neo4j_driver(neo4j_config)
    matched = 0
    updated = 0
    try:
        with driver.session() as session:
            for i in range(0, len(rows), int(batch_size)):
                batch = rows[i : i + int(batch_size)]
                if not batch:
                    continue
                res = session.run(
                    """
                    UNWIND $rows AS row
                    MATCH (m:Material)
                    WHERE elementId(m) = row.eid
                    WITH m, row
                    REMOVE m.image
                    WITH m, row
                    OPTIONAL MATCH (m)-[r:HAS_IMAGE]->(:MaterialImage)
                    DELETE r
                    WITH m, row
                    MERGE (img:MaterialImage {path: row.image})
                    SET img.source = 'xlsx',
                        img.updated_at = datetime(),
                        img.sheet = coalesce(row.sheet, img.sheet)
                    MERGE (m)-[rel:HAS_IMAGE]->(img)
                    SET rel.score = row.score,
                        rel.excel_code = row.excel_code,
                        rel.sheet = row.sheet,
                        rel.updated_at = datetime()
                    RETURN count(m) AS matched
                    """,
                    {"rows": batch},
                )
                for r in res:
                    if r and r.get("matched") is not None:
                        m = int(r["matched"])
                        matched += m
                        updated += m
    finally:
        try:
            driver.close()
        except Exception:
            pass
    return {"matched": matched, "updated": updated}


def update_material_images_as_nodes_by_code(
    *,
    rows: List[Dict[str, Any]],
    batch_size: int,
) -> Dict[str, int]:
    neo4j_config = get_neo4j_config()
    driver = get_neo4j_driver(neo4j_config)
    matched = 0
    updated = 0
    try:
        with driver.session() as session:
            for i in range(0, len(rows), int(batch_size)):
                batch = rows[i : i + int(batch_size)]
                if not batch:
                    continue
                res = session.run(
                    """
                    UNWIND $rows AS row
                    MATCH (m:Material {material_code: row.code})
                    WITH m, row
                    REMOVE m.image
                    WITH m, row
                    OPTIONAL MATCH (m)-[r:HAS_IMAGE]->(:MaterialImage)
                    DELETE r
                    WITH m, row
                    MERGE (img:MaterialImage {path: row.image})
                    SET img.source = 'xlsx',
                        img.updated_at = datetime(),
                        img.sheet = coalesce(row.sheet, img.sheet)
                    MERGE (m)-[rel:HAS_IMAGE]->(img)
                    SET rel.score = row.score,
                        rel.excel_code = row.excel_code,
                        rel.sheet = row.sheet,
                        rel.updated_at = datetime()
                    RETURN count(m) AS matched
                    """,
                    {"rows": batch},
                )
                for r in res:
                    if r and r.get("matched") is not None:
                        m = int(r["matched"])
                        matched += m
                        updated += m
    finally:
        try:
            driver.close()
        except Exception:
            pass
    return {"matched": matched, "updated": updated}


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _get_env(name: str, default: Optional[str] = None) -> Optional[str]:
    v = os.getenv(name)
    if v is None:
        return default
    v = v.strip()
    return v or default


def get_neo4j_config() -> Neo4jConfig:
    uri = _get_env("NEO4J_URI")
    user = _get_env("NEO4J_USER")
    password = _get_env("NEO4J_PASSWORD")
    if not uri or not user or not password:
        missing = [k for k, v in [("NEO4J_URI", uri), ("NEO4J_USER", user), ("NEO4J_PASSWORD", password)] if not v]
        raise RuntimeError(f"Missing Neo4j env vars: {', '.join(missing)}")
    return Neo4jConfig(uri=uri, user=user, password=password)


def detect_image_ext(data: bytes) -> str:
    if not data:
        return ".bin"
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"\xff\xd8"):
        return ".jpg"
    if data.startswith(b"GIF87a") or data.startswith(b"GIF89a"):
        return ".gif"
    if data.startswith(b"BM"):
        return ".bmp"
    return ".bin"


def sanitize_filename(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[\\/:*?\"<>|\s]+", "_", s)
    s = s.strip("._")
    return s or "unknown"


def _parse_col_spec(ws, col_spec: str, header_row: Optional[int]) -> int:
    spec = (col_spec or "").strip()
    if not spec:
        raise ValueError("empty material column spec")

    if spec.isdigit():
        idx = int(spec)
        if idx <= 0:
            raise ValueError("column index must be 1-based")
        return idx

    if re.fullmatch(r"[A-Za-z]+", spec):
        letters = spec.upper()
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx

    if header_row is None:
        raise ValueError("material-col is a header name but header_row is not set")

    target = _norm_header(spec)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        if _norm_header(str(v)) == target:
            return c

    raise ValueError(f"could not find header '{spec}' on row {header_row}")


def guess_header_row(ws, *, max_rows: int = 40) -> Optional[int]:
    candidates = {
        "model",
        "material_code",
        "material code",
        "物料编码",
        "物料",
        "编码",
    }
    for r in range(1, min(max_rows, ws.max_row) + 1):
        hits = 0
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if _norm_header(str(v)) in candidates:
                hits += 1
        if hits:
            return r
    return None


def _find_header_cols(ws, header_row: int, header_name: str) -> List[int]:
    target = _norm_header(header_name)
    cols: List[int] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        hv = _norm_header(str(v))
        if hv == target or (target and target in hv):
            cols.append(c)
    return cols


def _choose_nearest_header_col(cols: List[int], *, anchor_col: int) -> Optional[int]:
    if not cols:
        return None
    left = [c for c in cols if c <= anchor_col]
    if left:
        return max(left)
    right = [c for c in cols if c > anchor_col]
    if right:
        return min(right)
    return None


def list_sheets(xlsx_path: Path) -> List[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        try:
            wb.close()
        except Exception:
            pass


_NS = {
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "wb": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


def _zip_read_xml(zf: zipfile.ZipFile, name: str) -> ET.Element:
    data = zf.read(name)
    return ET.fromstring(data)


def _rels_map(zf: zipfile.ZipFile, rels_path: str) -> Dict[str, str]:
    try:
        root = _zip_read_xml(zf, rels_path)
    except KeyError:
        return {}
    out: Dict[str, str] = {}
    for rel in root.findall("rel:Relationship", _NS):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rid and target:
            out[rid] = target
    return out


def _resolve_target(base_dir: str, target: str) -> str:
    # base_dir like 'xl/worksheets' ; target like '../drawings/drawing1.xml'
    p = normpath(f"{base_dir}/{target}")
    # normpath may drop leading '..' resolution; ensure it doesn't start with '../'
    p = p.lstrip("./")
    return p


def _sheet_xml_path_from_name(zf: zipfile.ZipFile, sheet_name: str) -> Optional[str]:
    try:
        wb = _zip_read_xml(zf, "xl/workbook.xml")
    except KeyError:
        return None

    rid = None
    for sh in wb.findall("wb:sheets/wb:sheet", _NS):
        if sh.attrib.get("name") == sheet_name:
            rid = sh.attrib.get(f"{{{_NS['r']}}}id")
            break
    if not rid:
        return None

    rels = _rels_map(zf, "xl/_rels/workbook.xml.rels")
    target = rels.get(rid)
    if not target:
        return None
    return _resolve_target("xl", target)


def _iter_images_from_xlsx_zip(xlsx_path: Path, sheet_name: str) -> Iterable[Tuple[int, int, bytes]]:
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        sheet_xml = _sheet_xml_path_from_name(zf, sheet_name)
        if not sheet_xml:
            return

        try:
            sheet_root = _zip_read_xml(zf, sheet_xml)
        except KeyError:
            return

        # sheet rels: find drawings targets
        sheet_base_dir = str(Path(sheet_xml).parent).replace("\\", "/")
        sheet_rels = _rels_map(zf, f"{sheet_base_dir}/_rels/{Path(sheet_xml).name}.rels")

        drawing_targets: List[str] = []
        for dr in sheet_root.findall("wb:drawing", _NS):
            drid = dr.attrib.get(f"{{{_NS['r']}}}id")
            if not drid:
                continue
            t = sheet_rels.get(drid)
            if not t:
                continue
            drawing_targets.append(_resolve_target(sheet_base_dir, t))

        for drawing_xml in drawing_targets:
            try:
                drawing_root = _zip_read_xml(zf, drawing_xml)
            except KeyError:
                continue

            drawing_base_dir = str(Path(drawing_xml).parent).replace("\\", "/")
            drawing_rels = _rels_map(zf, f"{drawing_base_dir}/_rels/{Path(drawing_xml).name}.rels")

            anchors = drawing_root.findall("xdr:twoCellAnchor", _NS) + drawing_root.findall("xdr:oneCellAnchor", _NS)
            for anc in anchors:
                frm = anc.find("xdr:from", _NS)
                if frm is None:
                    continue
                row_el = frm.find("xdr:row", _NS)
                col_el = frm.find("xdr:col", _NS)
                if row_el is None or col_el is None:
                    continue
                try:
                    row1 = int(row_el.text or "0") + 1
                    col1 = int(col_el.text or "0") + 1
                except Exception:
                    continue

                blip = anc.find(".//a:blip", _NS)
                if blip is None:
                    continue
                rid = blip.attrib.get(f"{{{_NS['r']}}}embed")
                if not rid:
                    continue
                tgt = drawing_rels.get(rid)
                if not tgt:
                    continue
                media_path = _resolve_target(drawing_base_dir, tgt)
                try:
                    data = zf.read(media_path)
                except KeyError:
                    continue
                yield row1, col1, data


def iter_sheet_images(ws, *, xlsx_path: Path, sheet_name: str) -> Iterable[Tuple[int, int, object, Optional[bytes]]]:
    imgs = getattr(ws, "_images", None) or []
    if imgs:
        for img in imgs:
            anchor = getattr(img, "anchor", None)
            from_ = getattr(anchor, "_from", None)
            row0 = getattr(from_, "row", None)
            col0 = getattr(from_, "col", None)
            if row0 is None:
                continue
            try:
                row1 = int(row0) + 1
            except Exception:
                continue
            try:
                col1 = int(col0) + 1 if col0 is not None else 1
            except Exception:
                col1 = 1
            yield row1, col1, img, None
        return

    # Fallback: parse the XLSX zip and locate drawing anchors + media.
    for row1, col1, data in _iter_images_from_xlsx_zip(xlsx_path, sheet_name):
        yield row1, col1, None, data


def extract_images_and_map(
    *,
    ws,
    material_col: Optional[int],
    material_cols: Optional[List[int]],
    header_row: Optional[int],
    out_dir: Path,
    overwrite: bool,
    limit: Optional[int],
    scan_up: int,
    debug: bool,
    xlsx_path: Path,
    sheet_name: str,
) -> List[Tuple[str, Path]]:
    out_dir.mkdir(parents=True, exist_ok=True)

    results: List[Tuple[str, Path]] = []

    code_counts: Dict[str, int] = {}
    count = 0
    for row1, col1, img, raw in iter_sheet_images(ws, xlsx_path=xlsx_path, sheet_name=sheet_name):
        if header_row is not None and row1 == header_row:
            continue

        col_for_row = material_col
        if material_cols:
            col_for_row = _choose_nearest_header_col(material_cols, anchor_col=col1)
        if not col_for_row:
            continue

        col_for_row_int = int(col_for_row)
        code_val = ws.cell(row=row1, column=col_for_row_int).value
        code = str(code_val).strip() if code_val is not None else ""
        if not code and scan_up and header_row is not None:
            # Many Excel sheets store a model/material value in a merged cell.
            # openpyxl will only expose the value on the top-left cell, so for
            # images anchored within the merged region we need to look upward.
            lower_bound = max(1, (header_row + 1) if header_row else 1)
            for r in range(row1 - 1, max(lower_bound - 1, row1 - int(scan_up) - 1), -1):
                v = ws.cell(row=r, column=col_for_row_int).value
                if v is None:
                    continue
                candidate = str(v).strip()
                if candidate:
                    code = candidate
                    break
        if not code:
            if debug:
                print(f"[skip] row={row1} col={col1} material_col={col_for_row_int} reason=no_code")
            continue

        code = re.sub(r"\s+", " ", code).strip()

        data = None
        if raw is not None:
            data = raw
        else:
            try:
                data = img._data()  # openpyxl internal
            except Exception:
                data = None
        if not data:
            if debug:
                print(f"[skip] row={row1} col={col1} code={code} reason=no_image_data")
            continue
        ext = detect_image_ext(data)
        base = sanitize_filename(code)

        code_counts[base] = code_counts.get(base, 0) + 1
        idx = code_counts[base]
        name = f"{base}{'' if idx == 1 else f'_{idx}'}{ext}"
        dst = out_dir / name
        if dst.exists() and not overwrite:
            results.append((code, dst))
        else:
            dst.write_bytes(data)
            results.append((code, dst))

        if debug:
            print(f"[ok] row={row1} col={col1} code={code} -> {dst.name}")

        count += 1
        if limit is not None and count >= limit:
            return results

    return results


def update_material_images(
    *,
    mappings: List[Tuple[str, str]],
    batch_size: int,
) -> Dict[str, int]:
    if not mappings:
        return {"matched": 0, "updated": 0}

    neo4j_config = get_neo4j_config()
    driver = get_neo4j_driver(neo4j_config)

    matched = 0
    updated = 0
    try:
        with driver.session() as session:
            for i in range(0, len(mappings), batch_size):
                batch = mappings[i : i + batch_size]
                res = session.run(
                    """
                    UNWIND $rows AS row
                    MATCH (m:Material {material_code: row.code})
                    WITH m, row
                    SET m.image = row.image
                    RETURN count(m) AS matched
                    """,
                    {"rows": [{"code": c, "image": p} for c, p in batch]},
                )
                r = res.single()
                if r and r.get("matched") is not None:
                    m = int(r["matched"])
                    matched += m
                    updated += m
    finally:
        try:
            driver.close()
        except Exception:
            pass

    return {"matched": matched, "updated": updated}


def update_material_images_by_id(
    *,
    mappings: List[Tuple[int, str]],
    batch_size: int,
) -> Dict[str, int]:
    if not mappings:
        return {"matched": 0, "updated": 0}

    neo4j_config = get_neo4j_config()
    driver = get_neo4j_driver(neo4j_config)

    matched = 0
    updated = 0
    try:
        with driver.session() as session:
            for i in range(0, len(mappings), batch_size):
                batch = mappings[i : i + batch_size]
                res = session.run(
                    """
                    UNWIND $rows AS row
                    MATCH (m:Material)
                    WHERE id(m) = row.nid
                    WITH m, row
                    SET m.image = row.image
                    RETURN count(m) AS matched
                    """,
                    {"rows": [{"nid": nid, "image": image} for nid, image in batch]},
                )
                r = res.single()
                if r and r.get("matched") is not None:
                    m = int(r["matched"])
                    matched += m
                    updated += m
    finally:
        try:
            driver.close()
        except Exception:
            pass

    return {"matched": matched, "updated": updated}


def fuzzy_match_mappings(
    *,
    excel_mappings: List[Tuple[str, str, str]],
    candidates: List[Tuple[str, str]],
    min_score: float,
    top_k: int,
    ambiguous_delta: float,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    matched_updates: List[Dict[str, Any]] = []
    logs: List[str] = []

    if not excel_mappings:
        return matched_updates, logs

    if not candidates:
        for code, image, _sheet in excel_mappings:
            logs.append(f"SKIP\t{code}\tNO_CANDIDATES\t{image}")
        return matched_updates, logs

    top_k = max(1, int(top_k or 1))

    for excel_code, image, sheet in excel_mappings:
        scored: List[Tuple[float, str, str]] = []
        for eid, cand_code in candidates:
            score = _similarity(excel_code, cand_code)
            if score <= 0:
                continue
            scored.append((score, eid, cand_code))
        scored.sort(key=lambda x: x[0], reverse=True)

        if not scored:
            logs.append(f"SKIP\t{excel_code}\tNO_MATCH\t{image}")
            continue

        best_score, best_eid, best_code = scored[0]
        top = scored[:top_k]

        if best_score < float(min_score):
            logs.append(
                f"SKIP\t{excel_code}\tLOW_SCORE\tbest={best_code}\tscore={best_score:.3f}\t{image}"
            )
            continue

        if len(scored) >= 2:
            second_score, _, second_code = scored[1]
            if (best_score - second_score) < float(ambiguous_delta):
                logs.append(
                    f"SKIP\t{excel_code}\tAMBIGUOUS\tbest={best_code}({best_score:.3f})\tsecond={second_code}({second_score:.3f})\t{image}"
                )
                for s, _, c in top:
                    logs.append(f"  TOP\t{excel_code}\t{c}\t{s:.3f}")
                continue

        matched_updates.append(
            {
                "eid": best_eid,
                "image": image,
                "score": float(best_score),
                "excel_code": excel_code,
                "sheet": sheet,
            }
        )
        logs.append(
            f"OK\t{excel_code}\tMATCH={best_code}\tscore={best_score:.3f}\t{image}"
        )
        for s, _, c in top:
            logs.append(f"  TOP\t{excel_code}\t{c}\t{s:.3f}")

    return matched_updates, logs


def fuzzy_match_nodes_to_images(
    *,
    excel_mappings: List[Tuple[str, str, str]],
    candidates: List[Tuple[str, str]],
    min_score: float,
    top_k: int,
    ambiguous_delta: float,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    updates: List[Dict[str, Any]] = []
    logs: List[str] = []

    if not candidates:
        logs.append("SKIP_NODE\tNO_CANDIDATES")
        return updates, logs
    if not excel_mappings:
        logs.append("SKIP_NODE\tNO_IMAGES")
        return updates, logs

    top_k = max(1, int(top_k or 1))

    # Track top-K images per node while scanning.
    node_top: Dict[str, List[Tuple[float, int]]] = {}
    node_best_img: Dict[str, int] = {}
    node_best_score: Dict[str, float] = {}
    node_second_score: Dict[str, float] = {}
    node_code: Dict[str, str] = {eid: code for eid, code in candidates}

    # Iterate over images and update per-node best/second/topK.
    for img_idx, (excel_code, _image, _sheet) in enumerate(excel_mappings):
        for eid, cand_code in candidates:
            score = _similarity(cand_code, excel_code)
            if score <= 0:
                continue

            best = node_best_score.get(eid, 0.0)
            second = node_second_score.get(eid, 0.0)
            if score > best:
                node_second_score[eid] = best
                node_best_score[eid] = score
                node_best_img[eid] = img_idx
            elif score > second:
                node_second_score[eid] = score

            tops = node_top.get(eid)
            if tops is None:
                node_top[eid] = [(score, img_idx)]
            else:
                tops.append((score, img_idx))
                tops.sort(key=lambda x: x[0], reverse=True)
                if len(tops) > top_k:
                    del tops[top_k:]

    # Create one proposal per node.
    proposals: List[Tuple[float, str, int]] = []
    skipped_nodes: List[Tuple[str, str]] = []
    for eid, _code in candidates:
        if eid not in node_best_score:
            skipped_nodes.append((eid, "NO_MATCH"))
            continue
        best_score = node_best_score[eid]
        if best_score < float(min_score):
            skipped_nodes.append((eid, "LOW_SCORE"))
            continue
        second_score = node_second_score.get(eid, 0.0)
        if (best_score - second_score) < float(ambiguous_delta):
            skipped_nodes.append((eid, "AMBIGUOUS"))
            continue
        img_idx = node_best_img.get(eid)
        if img_idx is None:
            skipped_nodes.append((eid, "NO_IMAGE"))
            continue
        proposals.append((best_score, eid, img_idx))

    # Resolve conflicts by highest score globally; one image can be used once.
    proposals.sort(key=lambda x: x[0], reverse=True)
    used_images: Dict[int, Tuple[str, float]] = {}
    used_nodes: Dict[str, float] = {}
    for score, eid, img_idx in proposals:
        if eid in used_nodes:
            continue
        if img_idx in used_images:
            prev_nid, prev_score = used_images[img_idx]
            logs.append(
                f"SKIP_NODE\t{node_code.get(eid,'')}\tCONFLICT_IMAGE\timage_idx={img_idx}\tkept={node_code.get(prev_nid,'')}({prev_score:.3f})\tscore={score:.3f}"
            )
            continue
        used_images[img_idx] = (eid, score)
        used_nodes[eid] = score
        excel_code, image, sheet = excel_mappings[img_idx]
        logs.append(
            f"OK_NODE\t{node_code.get(eid,'')}\tPICK={excel_code}\tscore={score:.3f}\t{image}"
        )
        for s, ii in node_top.get(eid, [])[:top_k]:
            ec, _im, _sh = excel_mappings[ii]
            logs.append(f"  TOP_IMG\t{node_code.get(eid,'')}\t{ec}\t{s:.3f}")
        updates.append(
            {
                "eid": eid,
                "image": image,
                "score": float(score),
                "excel_code": excel_code,
                "sheet": sheet,
            }
        )

    for eid, reason in skipped_nodes:
        c = node_code.get(eid, "")
        b = node_best_score.get(eid, 0.0)
        s = node_second_score.get(eid, 0.0)
        logs.append(f"SKIP_NODE\t{c}\t{reason}\tbest={b:.3f}\tsecond={s:.3f}")
        for ss, ii in node_top.get(eid, [])[:top_k]:
            ec, _im, _sh = excel_mappings[ii]
            logs.append(f"  TOP_IMG\t{c}\t{ec}\t{ss:.3f}")

    return updates, logs


def _parse_sheets(sheet: str, sheets: str) -> List[str]:
    if sheets:
        out: List[str] = []
        for part in sheets.split(","):
            name = part.strip()
            if name:
                out.append(name)
        if out:
            return out
    if sheet:
        name = sheet.strip()
        if name:
            return [name]
    return list(DEFAULT_SHEETS)


def _build_mappings(
    *,
    pairs: List[Tuple[str, Path]],
    out_base: Path,
    path_prefix: str,
    sheet: str,
) -> List[Tuple[str, str, str]]:
    mappings: List[Tuple[str, str, str]] = []
    for code, path in pairs:
        if path_prefix:
            rel = path.relative_to(out_base)
            stored = str(Path(path_prefix) / rel)
        else:
            stored = str(path)
        mappings.append((code, stored, sheet))
    return mappings


def main() -> None:
    load_dotenv()

    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=DEFAULT_XLSX_PATH, help="Path to .xlsx")
    p.add_argument("--list-sheets", action="store_true", help="List sheet names and exit")
    p.add_argument("--sheet", default="", help="Sheet name to process")
    p.add_argument("--sheets", default="", help="Comma-separated sheet names to process")
    p.add_argument("--material-col", default="Model", help="Material column: header name OR column letter (e.g. 'A') OR 1-based index")
    p.add_argument("--header-row", type=int, default=3, help="Header row (1-based). 0 means auto-detect")
    p.add_argument("--out-dir", default=str(BACKEND_ROOT / "material_images"), help="Output directory for extracted images")
    p.add_argument(
        "--path-prefix",
        default="/static/material_images",
        help="Prefix to store into image path. Default stores a portable URL under /static/material_images. Use empty string to store absolute path",
    )
    p.add_argument("--dry-run", action="store_true", help="Do not write to Neo4j")
    p.add_argument("--write", action="store_true", help="Write updates into Neo4j (default is dry-run)")
    p.add_argument("--fuzzy", action="store_true", default=True, help="Use fuzzy matching against Neo4j Material.material_code")
    p.add_argument("--no-fuzzy", action="store_false", dest="fuzzy", help="Disable fuzzy matching and use exact match")
    p.add_argument(
        "--fuzzy-direction",
        choices=["node-to-image", "image-to-node"],
        default="node-to-image",
        help="Fuzzy match direction: 'node-to-image' assigns one best image per node; 'image-to-node' matches each image to a node",
    )
    p.add_argument("--min-score", type=float, default=0.85, help="Min similarity score for fuzzy match")
    p.add_argument("--top-k", type=int, default=3, help="Show top K candidates for fuzzy match")
    p.add_argument("--ambiguous-delta", type=float, default=0.03, help="Skip if top-1 and top-2 scores differ by less than this")
    p.add_argument("--overwrite", action="store_true", help="Overwrite existing image files")
    p.add_argument("--clean-out-dir", action="store_true", help="Delete existing extracted images in the output sheet directory before writing")
    p.add_argument("--limit", type=int, default=0, help="Limit number of extracted images")
    p.add_argument("--scan-up", type=int, default=50, help="If material code is empty on the image row, scan upward N rows in the same column")
    p.add_argument("--show-matched", action="store_true", help="Print matched material codes list (effective when --write)")
    p.add_argument("--show-unmatched", action="store_true", help="Print unmatched material codes list with reasons (effective when --write)")
    p.add_argument("--max-show", type=int, default=50, help="Max rows to print for matched/unmatched lists")
    p.add_argument("--debug", action="store_true", help="Print debug logs for image anchors and mapping decisions")
    p.add_argument("--batch-size", type=int, default=200, help="Neo4j batch size")

    args = p.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    if args.list_sheets:
        for name in list_sheets(xlsx_path):
            print(name)
        return

    sheet_names = _parse_sheets(args.sheet or "", args.sheets or "")
    effective_dry_run = bool(args.dry_run) or (not bool(args.write))

    wb = load_workbook(xlsx_path, data_only=True)
    try:
        out_base = Path(args.out_dir).expanduser().resolve()
        out_base.mkdir(parents=True, exist_ok=True)
        limit = int(args.limit) if int(args.limit or 0) > 0 else None

        mapping_by_code: Dict[str, str] = {}
        mapping_sheet_by_code: Dict[str, str] = {}
        excel_mappings_all: List[Tuple[str, str, str]] = []
        total_pairs = 0
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"sheet not found: {sheet_name}")
            ws = wb[sheet_name]

            header_row = int(args.header_row or 0)
            if header_row <= 0:
                header_row = guess_header_row(ws) or 1

            material_col_spec = (args.material_col or "").strip()
            if not material_col_spec:
                raise ValueError("--material-col is required")

            material_col: Optional[int] = None
            material_cols: Optional[List[int]] = None
            spec = material_col_spec
            if spec.isdigit():
                material_col = _parse_col_spec(ws, spec, header_row)
            elif re.fullmatch(r"[A-Za-z]+", spec):
                cols = _find_header_cols(ws, header_row, spec)
                if cols:
                    if len(cols) == 1:
                        material_col = cols[0]
                    else:
                        material_cols = cols
                else:
                    material_col = _parse_col_spec(ws, spec, header_row)
            else:
                cols = _find_header_cols(ws, header_row, spec)
                if not cols:
                    raise ValueError(f"could not find header '{spec}' on row {header_row}")
                if len(cols) == 1:
                    material_col = cols[0]
                else:
                    material_cols = cols

            out_dir = out_base / sanitize_filename(sheet_name)
            if bool(args.clean_out_dir) and out_dir.exists() and out_dir.is_dir():
                exts = {".png", ".jpg", ".jpeg", ".gif", ".bmp"}
                removed = 0
                for pth in out_dir.iterdir():
                    if not pth.is_file():
                        continue
                    if pth.suffix.lower() not in exts:
                        continue
                    try:
                        pth.unlink()
                        removed += 1
                    except Exception:
                        pass
                if removed:
                    print(f"Cleaned {removed} existing images in: {out_dir}")

            pairs = extract_images_and_map(
                ws=ws,
                material_col=material_col,
                material_cols=material_cols,
                header_row=header_row,
                out_dir=out_dir,
                overwrite=bool(args.overwrite),
                limit=limit,
                scan_up=int(args.scan_up or 0),
                debug=bool(args.debug),
                xlsx_path=xlsx_path,
                sheet_name=sheet_name,
            )

            if not pairs:
                print(f"{sheet_name}: No images extracted or no matching rows.")
                continue

            total_pairs += len(pairs)
            mappings = _build_mappings(pairs=pairs, out_base=out_base, path_prefix=args.path_prefix or "", sheet=sheet_name)
            excel_mappings_all.extend(mappings)
            for code, stored, sh in mappings:
                mapping_by_code[code] = stored
                mapping_sheet_by_code[code] = sh

            print(f"{sheet_name}: Extracted {len(pairs)} images into: {out_dir}")
            print(f"{sheet_name}: Prepared {len(mappings)} material_code -> image mappings")
            if effective_dry_run:
                for code, stored, _sh in mappings[: min(20, len(mappings))]:
                    print(f"{code}\t{stored}")

        if total_pairs == 0:
            print("No images extracted or no matching rows.")
            return

        if args.fuzzy:
            excel_mappings = excel_mappings_all
            candidates = load_all_material_codes()
            if (args.fuzzy_direction or "") == "image-to-node":
                updates, match_logs = fuzzy_match_mappings(
                    excel_mappings=excel_mappings,
                    candidates=candidates,
                    min_score=float(args.min_score),
                    top_k=int(args.top_k),
                    ambiguous_delta=float(args.ambiguous_delta),
                )
            else:
                updates, match_logs = fuzzy_match_nodes_to_images(
                    excel_mappings=excel_mappings,
                    candidates=candidates,
                    min_score=float(args.min_score),
                    top_k=int(args.top_k),
                    ambiguous_delta=float(args.ambiguous_delta),
                )
            if effective_dry_run or args.debug:
                for line in match_logs:
                    print(line)
            else:
                ok_cnt = sum(1 for l in match_logs if l.startswith("OK\t") or l.startswith("OK_NODE\t"))
                skip_cnt = sum(1 for l in match_logs if l.startswith("SKIP\t") or l.startswith("SKIP_NODE\t"))
                print(f"Fuzzy match summary: ok={ok_cnt} skip={skip_cnt} candidates={len(candidates)}")

            if effective_dry_run:
                print("Dry-run: skip Neo4j update (pass --write to apply)")
                return

            # By default, show matched/unmatched lists when actually writing.
            max_show = max(1, int(args.max_show or 1))
            want_lists = True
            show_matched = bool(args.show_matched) or want_lists
            show_unmatched = bool(args.show_unmatched) or want_lists

            if (args.fuzzy_direction or "") == "image-to-node":
                # image-to-node logs are image-centric; matched/unmatched are less meaningful for nodes.
                if show_unmatched:
                    print("Unmatched nodes list is only available in node-to-image mode.")
            else:
                cand_by_eid: Dict[str, str] = {eid: code for eid, code in candidates}
                matched_eids = {row.get("eid") for row in updates if row.get("eid")}
                matched_codes = [cand_by_eid.get(eid, str(eid)) for eid in sorted(matched_eids)]

                reasons: Dict[str, str] = {}
                for line in match_logs:
                    if not line.startswith("SKIP_NODE\t"):
                        continue
                    parts = line.split("\t")
                    if len(parts) >= 3:
                        code = parts[1]
                        reason = parts[2]
                        if code and code not in reasons:
                            reasons[code] = reason

                unmatched_codes: List[str] = []
                for eid, code in candidates:
                    if eid in matched_eids:
                        continue
                    unmatched_codes.append(code)

                if show_matched:
                    print(f"Matched material_code count={len(matched_codes)}")
                    for c in matched_codes[:max_show]:
                        print(f"MATCHED\t{c}")
                    if len(matched_codes) > max_show:
                        print(f"... ({len(matched_codes) - max_show} more)")

                if show_unmatched:
                    print(f"Unmatched material_code count={len(unmatched_codes)}")
                    for c in unmatched_codes[:max_show]:
                        print(f"UNMATCHED\t{c}\t{reasons.get(c, 'NO_MATCH')}")
                    if len(unmatched_codes) > max_show:
                        print(f"... ({len(unmatched_codes) - max_show} more)")

            stats = update_material_images_as_nodes_by_eid(rows=updates, batch_size=int(args.batch_size))
            print(f"Neo4j updated (fuzzy, nodes). matched={stats['matched']} updated={stats['updated']}")
        else:
            if effective_dry_run:
                print("Dry-run: skip Neo4j update (pass --write to apply)")
                return
            rows: List[Dict[str, Any]] = []
            for code, image in mapping_by_code.items():
                rows.append(
                    {
                        "code": code,
                        "image": image,
                        "score": 1.0,
                        "excel_code": code,
                        "sheet": mapping_sheet_by_code.get(code, ""),
                    }
                )
            stats = update_material_images_as_nodes_by_code(rows=rows, batch_size=int(args.batch_size))
            print(f"Neo4j updated (exact, nodes). matched={stats['matched']} updated={stats['updated']}")
    finally:
        try:
            wb.close()
        except Exception:
            pass


if __name__ == "__main__":
    main()
